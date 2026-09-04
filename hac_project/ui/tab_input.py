"""
TAB 1: INPUT — กรอกข้อมูล HAC
"""
import streamlit as st
import pandas as pd
import os
import openpyxl

from constants import SOURCE_OPTIONS
from engine.pairing import parse_rack_layout
from ui.svg_diagram import build_hac_svg

EXCEL_IMPORT_PATH = "Input_datahall.xlsx"  # วางไว้ที่ hac_project/ (ระดับเดียวกับ app.py)


def load_hac_df_from_excel(sheet_name: str) -> pd.DataFrame:
    """
    อ่าน sheet ที่เลือกจาก Input_datahall.xlsx → DataFrame รูปแบบเดียวกับ hac_df
    โครงสร้างไฟล์: แถว 1 = header, แถว 2 เป็นต้นไป = ข้อมูล
    คอลัมน์ A = HAC Name, B = Source type, C เป็นต้นไป = Layout Racks (ความยาวไม่เท่ากันได้)
    """
    wb = openpyxl.load_workbook(EXCEL_IMPORT_PATH, data_only=True)
    ws = wb[sheet_name]

    rows_out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        source = row[1]
        if name is None or source is None:
            continue  # แถวว่าง ข้ามไป
        rack_values = [v for v in row[2:] if v is not None]
        rack_str = ", ".join(str(int(v)) if float(v).is_integer() else str(v) for v in rack_values)
        rows_out.append({
            "HAC Name": str(name),
            "Rack Layout (kW)": rack_str,
            "Source Type": str(source),
        })

    return pd.DataFrame(rows_out)


def render():
    st.header("กรอกข้อมูล HAC")
    st.caption("แต่ละ HAC มี 2 แถว (บน/ล่าง) | Source Type: 2-source (default) หรือ 4-source สำหรับ Liquid rack ≥100 kW")

 # ── SECTION: Import จาก Excel ────────────────────────────
    with st.expander("📥 Import ข้อมูลจาก Excel", expanded=False):
        if os.path.exists(EXCEL_IMPORT_PATH):
            try:
                sheet_names = openpyxl.load_workbook(EXCEL_IMPORT_PATH, read_only=True).sheetnames
                col1, col2 = st.columns([3, 1])
                with col1:
                    selected_sheet = st.selectbox("เลือก Sheet", options=sheet_names, key="excel_sheet_select")
                with col2:
                    st.write(""); st.write("")
                    if st.button("📥 โหลดข้อมูล", use_container_width=True, type="primary"):
                        new_df = load_hac_df_from_excel(selected_sheet)
                        if len(new_df) == 0:
                            st.warning(f"⚠️ ไม่พบข้อมูลใน sheet '{selected_sheet}'")
                        else:
                            st.session_state.hac_df = new_df   # ทับตารางเดิมทั้งหมด
                            st.success(f"✅ โหลด {len(new_df)} แถวจาก sheet '{selected_sheet}' แล้ว")
                            st.rerun()
            except Exception as e:
                st.error(f"❌ อ่านไฟล์ไม่สำเร็จ: {e}")
        else:
            st.info(f"ℹ️ ไม่พบไฟล์ `{EXCEL_IMPORT_PATH}` — วางไฟล์ไว้ในโฟลเดอร์เดียวกับ app.py แล้ว rerun ใหม่")

    st.divider()

    # default data — Rack Layout แบบ string
    if "hac_df" not in st.session_state:
        st.session_state.hac_df = pd.DataFrame([
            {"HAC Name": "HAC 1", "Rack Layout (kW)": "20, 20, 20, 20, 20, 20, 20, 20, 20, 20", "Source Type": "2-source"},
            {"HAC Name": "HAC 2", "Rack Layout (kW)": "20, 20, 20, 20, 20, 20, 20, 20, 20, 20", "Source Type": "2-source"},
            {"HAC Name": "HAC 3", "Rack Layout (kW)": "20, 150, 150, 150, 150, 150, 150, 150, 20, 20", "Source Type": "2-source"},
            {"HAC Name": "HAC 4", "Rack Layout (kW)": "20, 150, 150, 150, 150, 150, 150, 150, 20, 20", "Source Type": "2-source"},
            {"HAC Name": "HAC 5", "Rack Layout (kW)": "20, 150, 150, 150, 150, 150, 150, 150, 20, 20", "Source Type": "2-source"},
        ])

    # migrate df เก่าที่ยังใช้ Rack Count + Load per Rack
    df_cols = st.session_state.hac_df.columns.tolist()
    if "Rack Layout (kW)" not in df_cols and "Rack Count" in df_cols:
        old = st.session_state.hac_df
        st.session_state.hac_df = pd.DataFrame([
            {
                "HAC Name":        r["HAC Name"],
                "Rack Layout (kW)": ", ".join([str(int(r["Load per Rack (kW)"]))] * int(r["Rack Count"])),
                "Source Type":     r.get("Source Type", "2-source"),
            }
            for _, r in old.iterrows()
        ])
    if "Source Type" not in st.session_state.hac_df.columns:
        st.session_state.hac_df["Source Type"] = "2-source"

    edited_df = st.data_editor(
        st.session_state.hac_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "HAC Name": st.column_config.TextColumn("HAC Name", required=True),
            "Rack Layout (kW)": st.column_config.TextColumn(
                "Rack Layout (kW)",
                required=True,
                help="กรอกขนาด kW ของแต่ละตู้คั่นด้วยจุลภาค เช่น: 20, 150, 150, 150, 20",
            ),
            "Source Type": st.column_config.SelectboxColumn(
                "Source Type",
                options=SOURCE_OPTIONS,
                required=True,
                help="2-source = Dual-cord | 4-source = รับจากทุก UPS",
            ),
        },
        key="hac_editor",
    )
    st.session_state.hac_df = edited_df

    # Duplicate button
    if len(edited_df) > 0:
        col1, col2 = st.columns([3, 1])
        with col1:
            dup_choice = st.selectbox("Duplicate HAC", options=edited_df["HAC Name"].tolist(), key="dup_select")
        with col2:
            st.write(""); st.write("")
            if st.button("➕ Duplicate", use_container_width=True):
                row     = edited_df[edited_df["HAC Name"] == dup_choice].iloc[0].copy()
                base    = row["HAC Name"]
                existing = set(edited_df["HAC Name"].tolist())
                new_name = base + " copy"
                c = 2
                while new_name in existing:
                    new_name = f"{base} copy{c}"; c += 1
                row["HAC Name"] = new_name
                st.session_state.hac_df = pd.concat(
                    [st.session_state.hac_df, pd.DataFrame([row])], ignore_index=True
                )
                st.rerun()

    # Summary metrics — parse rack layout
    edited_df = edited_df.dropna(subset=["HAC Name"]).copy()
    edited_df["_rack_list"] = edited_df["Rack Layout (kW)"].apply(parse_rack_layout)
    edited_df["_rack_count"] = edited_df["_rack_list"].apply(len)
    edited_df["_row_kw"]    = edited_df["_rack_list"].apply(sum)
    edited_df["Total kW / HAC"] = edited_df["_row_kw"] * 2

    # validate — แสดง error ถ้า parse ไม่ได้
    invalid = edited_df[edited_df["_rack_count"] == 0]["HAC Name"].tolist()
    if invalid:
        st.warning(f"⚠️ กรอก Rack Layout ไม่ถูกต้องใน: {', '.join(invalid)} — ตัวอย่าง: 20, 150, 150, 20")

    total_all   = edited_df["Total kW / HAC"].sum()
    cnt_2src    = (edited_df["Source Type"] == "2-source").sum()
    cnt_4src    = (edited_df["Source Type"] == "4-source").sum()

    st.divider()
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("จำนวน HAC",          len(edited_df))
    c2.metric("จำนวนแถวทั้งหมด",    len(edited_df) * 2)
    c3.metric("Total IT Load (kW)", f"{total_all:,.0f}")
    c4.metric("HAC แบบ 2-source",   cnt_2src)
    c5.metric("HAC แบบ 4-source",   cnt_4src)

    # Settings
    st.divider()
    st.subheader("⚙️ ตั้งค่า Optimization")
    n_groups = st.number_input("จำนวนกลุ่ม Generator (default = 3)", min_value=2, max_value=6, value=3, step=1)
    st.session_state.n_groups = int(n_groups)

    # Diagram
    st.divider()
    st.subheader("แผนภาพ Data Hall")
    if len(edited_df) > 0:
        hac_list = [
            {
                "name":        r["HAC Name"],
                "count":       r["_rack_count"],
                "load":        r["_row_kw"] / r["_rack_count"] if r["_rack_count"] > 0 else 0,
                "rack_list":   r["_rack_list"],
                "source_type": r.get("Source Type", "2-source"),
            }
            for _, r in edited_df.iterrows()
        ]
        st.markdown(build_hac_svg(hac_list), unsafe_allow_html=True)
        st.caption("ขอบสีม่วง = 4-source | ตัวเลขในแต่ละตู้คือ kW จริง | สีพื้นหลังแถวจะแสดงหลังคำนวณ")

    st.divider()
    if st.button("🚀 คำนวณ Pairing Optimization", type="primary", use_container_width=True):
        st.session_state.run_optimization = True
        st.success("✅ คำนวณเสร็จแล้ว — เปิดแท็บ ผลลัพธ์ เพื่อดูผล")
